import openpyxl,os,time
from openpyxl.utils import get_column_letter
wb_A = openpyxl.Workbook()
sheet_a = wb_A.get_sheet_by_name('Sheet')
output = 'output.xlsx'
arr_f = [f for f in os.listdir(os.getcwd()) if f.endswith(".xlsx") and (f != output) ]
length=len(arr_f)
loop = 0
for f in arr_f:
	print ('Begin load '+str(loop+1)+'/'+str(length)+' : '+f)
	wb_B = openpyxl.load_workbook(f)
	wb_B_sheets = wb_B.get_sheet_names()
	for S1 in wb_B_sheets:
		# print("\t"+"S1:"+S1) 
		if loop== 0:
			r_a = 0
		else:
			r_a = sheet_a.max_row
		c_a = sheet_a.max_column
		sheet_b = wb_B.get_sheet_by_name(S1)
		r_b = sheet_b.max_row
		c_b = sheet_b.max_column
		
		# print('\t\t'+'r_a:'+str(r_a)+'  c_a:'+str(c_a)+'  r_b:'+str(r_b)+'  c_b:'+str(c_b))
		if r_b == 1 and c_b == 1:
			continue	
			
		for row in range(1+(loop!=0), r_b+1 ):
			for col in range(1, c_b+1):
				sheet_a[get_column_letter(col)+str(r_a+row)].value=sheet_b[get_column_letter(col)+str(row)].value
		loop += 1 
	print ('End   load '+str(loop)+'/'+str(length)+' : '+f)
wb_A.save(output)
print('Done..Will finish after 5s')
time.sleep(5)