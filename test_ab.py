
#	pdb.set_trace()
#	sheet.max_column

#import openpyxl, pprint, pdb
import openpyxl,os,xlwt
# import openpyxl,os,xlwt
from openpyxl.utils import get_column_letter
# wb_A = xlwt.Workbook()
# wb_A = openpyxl.load_workbook('test_ab.xlsx')
arr_f = [f for f in os.listdir(os.getcwd()) if f.endswith(".xlsx")]
for f in arr_f:
	print ("Begin load:"+f)
	if f == arr_f[0]:
		wb_A = openpyxl.load_workbook(f)
		continue
	wb_B = openpyxl.load_workbook(f)

	sheet_a = wb_A.get_sheet_by_name('Sheet1')
	r_a = sheet_a.max_row
	c_a = sheet_a.max_column
	sheet_b = wb_B.get_sheet_by_name('Sheet1')
	r_b = sheet_b.max_row
	c_b = sheet_b.max_column

	print('\t'+'r_a:'+str(r_a)+'  r_b:'+str(r_b)+'  c_a:'+str(c_a)+'  c_b:'+str(c_b))

	
	for row in range(1, r_b+1 ):
		for col in range(1, c_b+1):
			sheet_a[get_column_letter(col)+str(r_a+row)].value=sheet_b[get_column_letter(col)+str(row)].value
		

	print ("End load:"+f)
print('Save to test_ab.xlsx...')
wb_A.save('test_ab.xlsx')
print('Done.')
